# apps/quality/views/qwall_views.py
import os
import io
import requests
from datetime import date, timedelta, datetime as dt
from django.core.cache import cache
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.quality.services.qwall_service import QWallService

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROXY_URL = os.getenv("QWALL_PaROXY_URL", "http://host.docker.internal:8002")
HEADERS = {"Authorization": f"Bearer {os.getenv('QWALL_PROXY_TOKEN', '')}"}

# Excel styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FAIL_FONT = Font(color="FF0000", bold=True)
BOLD_LG = Font(bold=True, size=12)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _get_locale(request) -> str:
    """Reusa el mecanismo ya existente en el proyecto: el frontend manda el
    idioma activo (i18next) en el header Accept-Language ("es"/"en")."""
    lang = request.headers.get("Accept-Language", "es")
    return (lang.split(",")[0].split("-")[0].strip().lower() or "es")


def _fmt_duration(seconds) -> str:
    if not seconds:
        return "0:00"
    s = int(seconds)
    return f"{s // 60}:{str(s % 60).zfill(2)}"


def _parse_bu_ids(request) -> list[int]:
    """
    Lee `bu_id` repetido (?bu_id=1&bu_id=2) del query string. Nombre distinto
    al `bu` generico del resto del proyecto a proposito: bu_id es un FK a
    ssi_BusinessUnits en CCS, un catalogo completamente independiente de
    BusinessUnit/BU_CHOICES (ssi_common) -- no es el mismo concepto.
    """
    raw = request.query_params.getlist("bu_id")
    return [int(v) for v in raw if str(v).strip()]


def _build_excel(data: dict, start_date: date, end_date: date, include_test: bool = False) -> bytes:
    wb = Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Resumen"

    summary = data["summary"]
    avg_dur = _fmt_duration(summary["avg_duration"])
    by_part = sorted(data.get("by_part", []), key=lambda x: x["total"], reverse=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    center = Alignment(horizontal="center", vertical="center")

    total_cols = 2 + len(by_part)

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 16
    for i in range(len(by_part)):
        col_letter = ws_sum.cell(row=1, column=3 + i).column_letter
        ws_sum.column_dimensions[col_letter].width = 16

    # Fila 1: Título
    titulo = "Reporte de Inspecciones — PRUEBAS" if include_test else "Reporte de Inspecciones de Calidad"
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws_sum.cell(row=1, column=1, value=titulo)
    t.font = Font(bold=True, size=12)
    t.fill = title_fill
    t.alignment = center
    t.border = border
    for c in range(2, total_cols + 1):
        ws_sum.cell(row=1, column=c).border = border
    ws_sum.row_dimensions[1].height = 22

    # Fila 2: Sub-headers
    ws_sum.row_dimensions[2].height = 18
    ws_sum.cell(row=2, column=1).border = border
    ws_sum.cell(row=2, column=1).fill = title_fill

    total_hdr = ws_sum.cell(row=2, column=2, value="Total")
    total_hdr.font = Font(bold=True, size=10)
    total_hdr.fill = title_fill
    total_hdr.alignment = center
    total_hdr.border = border

    for i, part in enumerate(by_part):
        cell = ws_sum.cell(row=2, column=3 + i, value=part["part_number"])
        cell.font = Font(bold=True, size=10)
        cell.fill = title_fill
        cell.alignment = center
        cell.border = border

    # Filas de datos
    rows_data = [
        ("Período", f"{start_date} a {end_date}", None, None),
        ("Generado", dt.now().strftime("%Y-%m-%d %H:%M:%S"), None, None),
        ("Total de Inspecciones", summary["total"], "total", None),
        ("PASS", summary["pass"], "pass", False),
        ("FAIL", summary["fail"], "fail", True),
        ("Tasa de Aprobación", f"{summary['pass_rate']:.2f}%", "pass_rate", None),
        ("Tiempo Promedio", avg_dur, None, None),
        ("Inspectores Únicos", summary["inspectors"], None, None),
        ("Part Numbers Únicos", summary["part_numbers"], None, None),
    ]

    for i, (label, value, part_field, is_fail_field) in enumerate(rows_data, start=3):
        ws_sum.row_dimensions[i].height = 18
        fill = title_fill if i % 2 == 0 else no_fill

        a = ws_sum.cell(row=i, column=1, value=label)
        a.font = Font(size=10)
        a.alignment = center
        a.border = border
        a.fill = fill

        b = ws_sum.cell(row=i, column=2, value=value)
        b.font = Font(size=10)
        b.alignment = center
        b.border = border
        b.fill = fill

        for j, part in enumerate(by_part):
            cell = ws_sum.cell(row=i, column=3 + j)
            cell.alignment = center
            cell.border = border
            cell.fill = fill

            if part_field and part_field in part:
                val = part[part_field]
                if part_field == "pass_rate":
                    val = f"{val:.1f}%"
                cell.value = val
                if is_fail_field and isinstance(val, (int, float)) and val > 0:
                    cell.font = Font(size=10, color="C00000", bold=True)
                else:
                    cell.font = Font(size=10)
            else:
                cell.value = ""
                cell.font = Font(size=10)

    # ── Hoja 2: Inspecciones ──────────────────────────────────────────────────
    ws = wb.create_sheet("Inspecciones")

    columns = [
        ("Fecha", "inspection_date", 14),
        ("Semana", "week_number", 10),
        ("Mes", "month_name", 14),
        ("Hora Inicio", "time_start", 12),
        ("Hora Fin", "time_end", 12),
        ("Duración", "_duration", 12),
        ("Inspector", "inspector", 22),
        ("Tipo", "inspection_type", 16),
        ("Resultado", "result", 12),
        ("QTY", "_qty", 8),
        ("WO", "work_order", 14),
        ("Part Number", "part_number", 20),
        ("Serial SSI", "serial_ssi", 20),
        ("Serial Client", "serial_volvo", 20),
        ("Fallas", "fail_modes", 40),
    ]

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    rows = data.get("rows", [])
    for row_idx, r in enumerate(rows, start=2):
        values = []
        for _, field, _ in columns:
            if field == "_duration":
                values.append(_fmt_duration(r.get("duration_seconds")))
            elif field == "_qty":
                values.append(1)
            else:
                values.append(r.get(field, ""))

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if columns[col_idx - 1][0] == "Resultado":
                if val == "FAIL":
                    cell.font = FAIL_FONT
            if r.get("result") == "FAIL":
                cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")

    # ── Pivot de fallas ───────────────────────────────────────────────────────
    fail_modes = data.get("fail_modes", [])
    if fail_modes:
        pivot_col = len(columns) + 2

        title_cell = ws.cell(row=1, column=pivot_col, value="ANÁLISIS DE FALLAS")
        title_cell.font = BOLD_LG

        hdr_falla = ws.cell(row=2, column=pivot_col, value="Falla")
        hdr_cant = ws.cell(row=2, column=pivot_col + 1, value="Cantidad")
        for cell in [hdr_falla, hdr_cant]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        piv_col_letter = ws.cell(row=1, column=pivot_col).column_letter
        piv_col_next_letter = ws.cell(row=1, column=pivot_col + 1).column_letter
        ws.column_dimensions[piv_col_letter].width = 50
        ws.column_dimensions[piv_col_next_letter].width = 12

        for i, fm in enumerate(fail_modes, start=3):
            ws.cell(row=i, column=pivot_col, value=f"{fm['code']} — {fm['name']}")
            ws.cell(row=i, column=pivot_col + 1, value=fm["count"])

        total_row = len(fail_modes) + 3
        total_cell = ws.cell(row=total_row, column=pivot_col, value="TOTAL")
        total_cell.font = BOLD_LG
        sum_cell = ws.cell(
            row=total_row,
            column=pivot_col + 1,
            value=f"=SUM({piv_col_next_letter}3:{piv_col_next_letter}{total_row - 1})",
        )
        sum_cell.font = BOLD_LG
        sum_cell.alignment = CENTER

    # ── Output ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class QWallReportView(APIView):
    """
    Reporte de inspecciones Q-Wall.

    `bu_id` acepta cero, uno o varios valores repetidos (?bu_id=1&bu_id=2) --
    contrato multi-select, igual que el resto de filtros del proyecto, pero
    con nombre propio porque bu_id es un FK a un catalogo de CCS
    (ssi_BusinessUnits) independiente de BU_CHOICES/BusinessUnit. Sin
    seleccion, QWallService.get_report trae todas las BU sin filtrar.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        start_raw = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw = request.query_params.get("end_date", str(today))
        fmt = request.query_params.get("export", "json")
        include_test = request.query_params.get("include_test", "false").lower() == "true"
        locale = _get_locale(request)

        try:
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        try:
            bu_ids = _parse_bu_ids(request)
        except ValueError:
            return Response({"error": "bu_id inválido."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        if (end_date - start_date).days > 180:
            return Response({"error": "Rango máximo permitido: 180 días."}, status=400)

        # ═══════════════════════════════════════════════════════════════════════
        # IMPORTANTE: QWallService.get_report es ESTÁTICO
        # ═══════════════════════════════════════════════════════════════════════
        data = QWallService.get_report(
            start_date, end_date, include_test=include_test,
            bu_ids=bu_ids or None, locale=locale,
        )

        if fmt == "xlsx":
            try:
                xlsx = _build_excel(data, start_date, end_date, include_test=include_test)
                filename = f"qwall_{start_raw}_{end_raw}{'_test' if include_test else ''}.xlsx"
                response = HttpResponse(
                    xlsx,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            except Exception as e:
                import traceback
                return Response({"error": str(e), "trace": traceback.format_exc()}, status=500)

        return Response(data)


class QWallTrendView(APIView):
    """Tendencia de Pass Rate (diaria/semanal/mensual) contra el target vigente."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        granularity = request.query_params.get("granularity", "daily")
        if granularity not in ("daily", "weekly", "monthly"):
            return Response({"error": "granularity debe ser daily, weekly o monthly."}, status=400)

        start_raw = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw = request.query_params.get("end_date", str(today))
        include_test = request.query_params.get("include_test", "false").lower() == "true"
        bu_raw = request.query_params.get("bu_id")

        try:
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        try:
            bu_id = int(bu_raw) if bu_raw else None
        except ValueError:
            return Response({"error": "bu_id inválido."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        data = QWallService.get_trend(
            start_date, end_date, granularity=granularity, include_test=include_test, bu_id=bu_id,
        )
        return Response(data)


class QWallParetoView(APIView):
    """Top Fail Modes (80/20) — un solo ranking agregado sobre la ventana de
    la granularidad pedida (no es una serie temporal)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        granularity = request.query_params.get("granularity", "daily")
        if granularity not in ("daily", "weekly", "monthly"):
            return Response({"error": "granularity debe ser daily, weekly o monthly."}, status=400)

        start_raw = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw = request.query_params.get("end_date", str(today))
        include_test = request.query_params.get("include_test", "false").lower() == "true"
        bu_raw = request.query_params.get("bu_id")
        limit_raw = request.query_params.get("limit", "10")
        locale = _get_locale(request)

        try:
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        try:
            bu_id = int(bu_raw) if bu_raw else None
        except ValueError:
            return Response({"error": "bu_id inválido."}, status=400)

        try:
            limit = int(limit_raw)
            if limit < 1 or limit > 50:
                raise ValueError
        except ValueError:
            return Response({"error": "limit inválido (1-50)."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        data = QWallService.get_pareto(
            start_date, end_date, granularity=granularity, include_test=include_test,
            bu_id=bu_id, locale=locale, limit=limit,
        )
        return Response(data)


class QWallFailByPointView(APIView):
    """Distribución de FALLAS entre puntos de inspección (de las piezas que
    fallaron, en qué punto ocurrió cada falla) — no % de fail por punto."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        start_raw = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw = request.query_params.get("end_date", str(today))
        include_test = request.query_params.get("include_test", "false").lower() == "true"
        bu_raw = request.query_params.get("bu_id")

        try:
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        try:
            bu_id = int(bu_raw) if bu_raw else None
        except ValueError:
            return Response({"error": "bu_id inválido."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        data = QWallService.get_fail_by_point(
            start_date, end_date, include_test=include_test, bu_id=bu_id,
        )
        return Response(data)


class QWallBuSummaryView(APIView):
    """Resumen agregado por Business Unit — usado por el dashboard cuando el
    filtro general está en 'Todas las BU' (Modo A del Part Number Summary)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        start_raw = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw = request.query_params.get("end_date", str(today))
        include_test = request.query_params.get("include_test", "false").lower() == "true"

        try:
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        data = QWallService.get_bu_summary(start_date, end_date, include_test=include_test)
        return Response(data)


class QWallPartNumberSummaryView(APIView):
    """Resumen de Part Numbers para UNA Business Unit específica. business_unit_id
    es obligatorio — este endpoint nunca regresa part numbers de todas las BU
    mezclados. Usado tanto en Modo B (filtro general ya en una BU) como al
    expandir una barra de BU en Modo A."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()
        start_raw = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw = request.query_params.get("end_date", str(today))
        include_test = request.query_params.get("include_test", "false").lower() == "true"
        bu_raw = request.query_params.get("business_unit_id")

        try:
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        if not bu_raw:
            return Response({"error": "business_unit_id es obligatorio."}, status=400)
        try:
            business_unit_id = int(bu_raw)
        except ValueError:
            return Response({"error": "business_unit_id inválido."}, status=400)

        data = QWallService.get_part_number_summary(
            start_date, end_date, business_unit_id, include_test=include_test,
        )
        return Response(data)


class QWallPartNumbersView(APIView):
    """Catálogo de part numbers desde Q-Wall (vía qwall-proxy)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cache_key = "qwall:part_numbers_catalog"
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)
        try:
            resp = requests.get(
                f"{PROXY_URL}/part-numbers",
                headers=HEADERS,
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json().get("data", [])
            cache.set(cache_key, data, 3600)
            return Response(data)
        except Exception as e:
            return Response({"detail": str(e)}, status=502)