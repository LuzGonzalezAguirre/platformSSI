from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .models import QualityTarget
from .serializers import QualityTargetSerializer
from .services.quality_service import QualityService
from apps.warehouse.services.plex_client import PlexProxyError
from datetime import date, timedelta
from apps.quality.services.qwall_service import QWallService

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta, datetime as dt

from .services.rejection_service import RejectionService

from .services.rejection_pdf_service import build_rejection_pdf
from .repositories.rejection_repository import RejectionRepository

import os
import requests
from django.core.cache import cache
from datetime import date, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from apps.quality.services.qwall_service import QWallService
PROXY_URL  = os.getenv("QWALL_PROXY_URL",   "http://host.docker.internal:8002")
HEADERS    = {"Authorization": f"Bearer {os.getenv('QWALL_PROXY_TOKEN', '')}"}

class ScrapDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start     = request.query_params.get("start_date")
        end       = request.query_params.get("end_date")
        use_shift = request.query_params.get("use_shift", "true").lower() == "true"

        if not start or not end:
            return Response({"detail": "start_date y end_date requeridos."}, status=400)
        try:
            data = QualityService().get_scrap_detail(start, end, use_shift)
            return Response(data)
        except PlexProxyError as e:
            return Response({"detail": str(e)}, status=502)


class QualityTargetView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        targets = QualityTarget.objects.all().order_by("level", "bu", "workcenter_name")
        return Response(QualityTargetSerializer(targets, many=True).data)

    def post(self, request):
        serializer = QualityTargetSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        serializer.save(updated_by=request.user)
        return Response(serializer.data, status=201)

    def put(self, request, pk: int):
        try:
            target = QualityTarget.objects.get(pk=pk)
        except QualityTarget.DoesNotExist:
            return Response({"detail": "No encontrado."}, status=404)
        serializer = QualityTargetSerializer(target, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        serializer.save(updated_by=request.user)
        return Response(serializer.data)

    def delete(self, request, pk: int):
        try:
            QualityTarget.objects.get(pk=pk).delete()
        except QualityTarget.DoesNotExist:
            return Response({"detail": "No encontrado."}, status=404)
        return Response(status=204)
    



class QWallReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today        = date.today()
        start_raw    = request.query_params.get("start_date", str(today - timedelta(days=7)))
        end_raw      = request.query_params.get("end_date",   str(today))
        fmt          = request.query_params.get("export",       "json")
        include_test = request.query_params.get("include_test", "false").lower() == "true"

        try:
            start_date = date.fromisoformat(start_raw)
            end_date   = date.fromisoformat(end_raw)
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=400)

        if end_date < start_date:
            return Response({"error": "end_date no puede ser anterior a start_date."}, status=400)

        if (end_date - start_date).days > 180:
            return Response({"error": "Rango máximo permitido: 180 días."}, status=400)

        data = QWallService.get_report(start_date, end_date, include_test=include_test)

        if fmt == "xlsx":
            try:
                xlsx     = _build_excel(data, start_date, end_date, include_test=include_test)
                filename = f"qwall_{start_raw}_{end_raw}{'_test' if include_test else ''}.xlsx"
                response = HttpResponse(
                    xlsx,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            except Exception as e:
                import traceback
                return Response({"error": str(e), "trace": traceback.format_exc()}, status=500)

        return Response(data)

# ── Excel builder ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FAIL_FONT   = Font(color="FF0000", bold=True)
BOLD_LG     = Font(bold=True, size=12)
BOLD        = Font(bold=True)
CENTER      = Alignment(horizontal="center", vertical="center")


def _fmt_duration(seconds) -> str:
    if not seconds:
        return "0:00"
    s = int(seconds)
    return f"{s // 60}:{str(s % 60).zfill(2)}"


def _build_excel(data: dict, start_date: date, end_date: date, include_test: bool = False) -> bytes:
    wb = Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Resumen"

    summary = data["summary"]
    avg_dur = _fmt_duration(summary["avg_duration"])
    by_part = sorted(data.get("by_part", []), key=lambda x: x["total"], reverse=True)

    thin       = Side(style="thin", color="000000")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    no_fill    = PatternFill(fill_type=None)
    center     = Alignment(horizontal="center", vertical="center")

    total_cols = 2 + len(by_part)

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 16
    for i in range(len(by_part)):
        col_letter = ws_sum.cell(row=1, column=3 + i).column_letter
        ws_sum.column_dimensions[col_letter].width = 16

    # Fila 1: Título
    titulo = "Reporte de Inspecciones — PRUEBAS" if include_test else "Reporte de Inspecciones de Calidad"
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t           = ws_sum.cell(row=1, column=1, value=titulo)
    t.font      = Font(bold=True, size=12)
    t.fill      = title_fill
    t.alignment = center
    t.border    = border
    for c in range(2, total_cols + 1):
        ws_sum.cell(row=1, column=c).border = border
    ws_sum.row_dimensions[1].height = 22

    # Fila 2: Sub-headers
    ws_sum.row_dimensions[2].height = 18
    ws_sum.cell(row=2, column=1).border = border
    ws_sum.cell(row=2, column=1).fill   = title_fill

    total_hdr           = ws_sum.cell(row=2, column=2, value="Total")
    total_hdr.font      = Font(bold=True, size=10)
    total_hdr.fill      = title_fill
    total_hdr.alignment = center
    total_hdr.border    = border

    for i, part in enumerate(by_part):
        cell           = ws_sum.cell(row=2, column=3 + i, value=part["part_number"])
        cell.font      = Font(bold=True, size=10)
        cell.fill      = title_fill
        cell.alignment = center
        cell.border    = border

    # Filas de datos
    rows_data = [
        ("Período",               f"{start_date} a {end_date}",          None,        None),
        ("Generado",              dt.now().strftime("%Y-%m-%d %H:%M:%S"), None,        None),
        ("Total de Inspecciones", summary["total"],                       "total",     None),
        ("PASS",                  summary["pass"],                        "pass",      False),
        ("FAIL",                  summary["fail"],                        "fail",      True),
        ("Tasa de Aprobación",    f"{summary['pass_rate']:.2f}%",         "pass_rate", None),
        ("Tiempo Promedio",       avg_dur,                                None,        None),
        ("Inspectores Únicos",    summary["inspectors"],                  None,        None),
        ("Part Numbers Únicos",   summary["part_numbers"],                None,        None),
    ]

    for i, (label, value, part_field, is_fail_field) in enumerate(rows_data, start=3):
        ws_sum.row_dimensions[i].height = 18
        fill = title_fill if i % 2 == 0 else no_fill

        a           = ws_sum.cell(row=i, column=1, value=label)
        a.font      = Font(size=10)
        a.alignment = center
        a.border    = border
        a.fill      = fill

        b           = ws_sum.cell(row=i, column=2, value=value)
        b.font      = Font(size=10)
        b.alignment = center
        b.border    = border
        b.fill      = fill

        for j, part in enumerate(by_part):
            cell           = ws_sum.cell(row=i, column=3 + j)
            cell.alignment = center
            cell.border    = border
            cell.fill      = fill

            if part_field and part_field in part:
                val = part[part_field]
                if part_field == "pass_rate":
                    val = f"{val:.1f}%"
                cell.value = val
                if is_fail_field and isinstance(val, (int, float)) and val > 0:
                    cell.font = Font(size=10, color="C00000", bold=True)
                else:
                    cell.font = Font(size=10)
            else:
                cell.value = ""
                cell.font  = Font(size=10)

    # ── Hoja 2: Inspecciones ──────────────────────────────────────────────────
    ws = wb.create_sheet("Inspecciones")

    columns = [
        ("Fecha",        "inspection_date", 14),
        ("Semana",       "week_number",     10),
        ("Mes",          "month_name",      14),
        ("Hora Inicio",  "time_start",      12),
        ("Hora Fin",     "time_end",        12),
        ("Duración",     "_duration",       12),
        ("Inspector",    "inspector",       22),
        ("Tipo",         "inspection_type", 16),
        ("Resultado",    "result",          12),
        ("QTY",          "_qty",            8),
        ("WO",           "work_order",      14),
        ("Part Number",  "part_number",     20),
        ("Serial SSI",   "serial_ssi",      20),
        ("Serial Volvo", "serial_volvo",    20),
        ("Fallas",       "fail_modes",      40),
    ]

    for col_idx, (header, _, width) in enumerate(columns, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.fill       = HEADER_FILL
        cell.font       = HEADER_FONT
        cell.alignment  = CENTER
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    rows = data.get("rows", [])
    for row_idx, r in enumerate(rows, start=2):
        values = []
        for _, field, _ in columns:
            if field == "_duration":
                values.append(_fmt_duration(r.get("duration_seconds")))
            elif field == "_qty":
                values.append(1)
            else:
                values.append(r.get(field, ""))

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if columns[col_idx - 1][0] == "Resultado":
                if val == "FAIL":
                    cell.font = FAIL_FONT
            if r.get("result") == "FAIL":
                cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")

    # ── Pivot de fallas ───────────────────────────────────────────────────────
    fail_modes = data.get("fail_modes", [])
    if fail_modes:
        pivot_col = len(columns) + 2

        title_cell      = ws.cell(row=1, column=pivot_col, value="ANÁLISIS DE FALLAS")
        title_cell.font = BOLD_LG

        hdr_falla = ws.cell(row=2, column=pivot_col,     value="Falla")
        hdr_cant  = ws.cell(row=2, column=pivot_col + 1, value="Cantidad")
        for cell in [hdr_falla, hdr_cant]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER

        piv_col_letter      = ws.cell(row=1, column=pivot_col).column_letter
        piv_col_next_letter = ws.cell(row=1, column=pivot_col + 1).column_letter
        ws.column_dimensions[piv_col_letter].width      = 50
        ws.column_dimensions[piv_col_next_letter].width = 12

        for i, fm in enumerate(fail_modes, start=3):
            ws.cell(row=i, column=pivot_col,     value=fm["fail_mode"])
            ws.cell(row=i, column=pivot_col + 1, value=fm["count"])

        total_row  = len(fail_modes) + 3
        total_cell = ws.cell(row=total_row, column=pivot_col, value="TOTAL")
        total_cell.font = BOLD_LG
        sum_cell = ws.cell(
            row=total_row, column=pivot_col + 1,
            value=f"=SUM({piv_col_next_letter}3:{piv_col_next_letter}{total_row - 1})"
        )
        sum_cell.font      = BOLD_LG
        sum_cell.alignment = CENTER

    # ── Output ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

class RejectionReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start        = request.query_params.get("start_date")
        end          = request.query_params.get("end_date")
        bu_id        = request.query_params.get("bu_id")
        include_test = request.query_params.get("include_test", "false").lower() == "true"

        if not start or not end:
            return Response({"detail": "start_date y end_date requeridos."}, status=400)

        try:
            data = RejectionService().get_tree(
                start, end, int(bu_id) if bu_id else None, include_test=include_test
            )
            return Response(data)
        except Exception as e:
            return Response({"detail": str(e)}, status=502)

class RejectionPhotoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, inspection_id: int):
        try:
            data = RejectionService().get_photo(inspection_id)
            return Response(data)
        except Exception as e:
            return Response({"detail": str(e)}, status=502)
        
class RejectionReportPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start = request.query_params.get("start_date")
        end   = request.query_params.get("end_date")
        bu_id = request.query_params.get("bu_id")
        lang  = request.query_params.get("lang", "es")   # ← agregar
        if lang not in ("es", "en"):
            lang = "es"

        if not start or not end:
            return Response({"detail": "start_date y end_date requeridos."}, status=400)

        try:
            svc  = RejectionService()
            tree = svc.get_tree(start, end, int(bu_id) if bu_id else None)

            repo = RejectionRepository()
            for node in tree:
                for serial in node["serials"]:
                    for insp in serial["inspections"]:
                        if insp.get("has_photo"):
                            try:
                                photo = repo.get_rejection_photo(insp["inspection_id"])
                                insp["photo_b64"] = photo.get("photo_b64")
                            except Exception:
                                insp["photo_b64"] = None
                        else:
                            insp["photo_b64"] = None

            pdf_bytes = build_rejection_pdf(tree, start, end, lang)   # ← pasar lang

            filename = f"rechazos_{start}_{end}.pdf"
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response({"detail": str(e)}, status=502)
        
# apps/backend/apps/quality/views.py — agregar

class QWallPartNumbersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cache_key = "qwall:part_numbers_catalog"
        cached    = cache.get(cache_key)
        if cached:
            return Response(cached)
        try:
            resp = requests.get(
                f"{PROXY_URL}/part-numbers",
                headers=HEADERS,
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json().get("data", [])
            cache.set(cache_key, data, 3600)  # 1 hora — catálogo estático
            return Response(data)
        except Exception as e:
            return Response({"detail": str(e)}, status=502) 