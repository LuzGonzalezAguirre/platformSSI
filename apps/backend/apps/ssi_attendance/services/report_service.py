"""
PDF and Excel report generator for attendance module.
"""
from datetime import date
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
)
from reportlab.lib.enums import TA_CENTER

from apps.ssi_common.db import execute_query

BRAND_BLUE = colors.HexColor("#0070C0")
BRAND_GREEN = colors.HexColor("#00B050")
LIGHT_GRAY = colors.HexColor("#F2F2F2")
MID_GRAY = colors.HexColor("#BFBFBF")


def _fetch_attendance(start_date: date, end_date: date, turno, department, employee_id) -> list[dict]:
    filters = ["a.attendance_date BETWEEN ? AND ?"]
    params: list = [start_date, end_date]
    if turno:
        filters.append("a.turno = ?")
        params.append(turno)
    if department:
        filters.append("e.department = ?")
        params.append(department)
    if employee_id:
        filters.append("a.employee_id = ?")
        params.append(employee_id)

    return execute_query(
        f"""
        SELECT
            a.attendance_date, a.check_in, a.check_out,
            a.turno, a.regular_hours, a.overtime_hours, a.total_hours, a.status,
            e.name AS employee_name, e.barcode_id, e.department
        FROM ssi_Attendance a
        INNER JOIN ssi_production_employee e ON a.employee_id = e.id
        WHERE {" AND ".join(filters)}
        ORDER BY a.attendance_date, e.name
        """,
        tuple(params),
    )


def generate_attendance_pdf(
    start_date: date,
    end_date: date,
    turno=None,
    department=None,
    employee_id=None,
    generated_by: str = "Sistema",
) -> bytes:
    rows = _fetch_attendance(start_date, end_date, turno, department, employee_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor=BRAND_BLUE, fontSize=13, spaceAfter=4)
    sub_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    small = ParagraphStyle("sm", parent=styles["Normal"], fontSize=7)
    hc = ParagraphStyle("hc", parent=styles["Normal"], fontSize=7, textColor=colors.white, alignment=TA_CENTER)

    story = []
    story.append(Paragraph("REPORTE DE ASISTENCIA - SSI Producción", title_style))
    story.append(Paragraph(f"Período: {start_date} al {end_date}   |   Generado: {date.today()}   |   Por: {generated_by}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=BRAND_BLUE, spaceAfter=8))

    headers = [
        Paragraph(h, hc) for h in
        ["Fecha", "Empleado", "Código", "Turno", "Entrada", "Salida",
         "H. Regulares", "H. Extras", "Total H.", "Estado"]
    ]
    col_widths = [0.85*inch, 2.0*inch, 1.0*inch, 0.6*inch, 0.9*inch, 0.9*inch,
                  0.9*inch, 0.8*inch, 0.8*inch, 0.85*inch]

    data = [headers]
    for r in rows:
        data.append([
            str(r.get("attendance_date") or ""),
            Paragraph(r.get("employee_name") or "", small),
            r.get("barcode_id") or "",
            r.get("turno") or "",
            r["check_in"].strftime("%H:%M") if r.get("check_in") else "-",
            r["check_out"].strftime("%H:%M") if r.get("check_out") else "-",
            str(r.get("regular_hours") or "-"),
            str(r.get("overtime_hours") or "-"),
            str(r.get("total_hours") or "-"),
            r.get("status") or "",
        ])

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ("GRID", (0, 0), (-1, -1), 0.3, MID_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t)

    def _page_num(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(landscape(letter)[0] - 0.5*inch, 0.35*inch, f"Página {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_page_num, onLaterPages=_page_num)
    return buffer.getvalue()


def generate_attendance_excel(
    start_date: date,
    end_date: date,
    turno=None,
    department=None,
    employee_id=None,
) -> bytes:
    """Generate Excel using openpyxl (if available) or CSV fallback."""
    rows = _fetch_attendance(start_date, end_date, turno, department, employee_id)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        header_fill = PatternFill("solid", fgColor="0070C0")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        headers = ["Fecha", "Empleado", "Código", "Turno", "Entrada", "Salida",
                   "H. Regulares", "H. Extras", "Total H.", "Estado", "Departamento"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=str(r.get("attendance_date") or ""))
            ws.cell(row=row_idx, column=2, value=r.get("employee_name") or "")
            ws.cell(row=row_idx, column=3, value=r.get("barcode_id") or "")
            ws.cell(row=row_idx, column=4, value=r.get("turno") or "")
            ws.cell(row=row_idx, column=5, value=r["check_in"].strftime("%H:%M") if r.get("check_in") else "")
            ws.cell(row=row_idx, column=6, value=r["check_out"].strftime("%H:%M") if r.get("check_out") else "")
            ws.cell(row=row_idx, column=7, value=float(r.get("regular_hours") or 0))
            ws.cell(row=row_idx, column=8, value=float(r.get("overtime_hours") or 0))
            ws.cell(row=row_idx, column=9, value=float(r.get("total_hours") or 0))
            ws.cell(row=row_idx, column=10, value=r.get("status") or "")
            ws.cell(row=row_idx, column=11, value=r.get("department") or "")

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except ImportError:
        # CSV fallback
        import csv
        buf = BytesIO()
        writer = csv.writer(buf)
        writer.writerow(["Fecha", "Empleado", "Código", "Turno", "Entrada", "Salida",
                         "H. Regulares", "H. Extras", "Total H.", "Estado"])
        for r in rows:
            writer.writerow([
                r.get("attendance_date"), r.get("employee_name"), r.get("barcode_id"),
                r.get("turno"),
                r["check_in"].strftime("%H:%M") if r.get("check_in") else "",
                r["check_out"].strftime("%H:%M") if r.get("check_out") else "",
                r.get("regular_hours"), r.get("overtime_hours"), r.get("total_hours"),
                r.get("status"),
            ])
        return buf.getvalue()
